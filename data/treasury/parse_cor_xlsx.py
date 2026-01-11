#!/usr/bin/env python3
"""
Parse COR-Summary_Monthly_1986-2024.xlsx and output a flat CSV
suitable for database import.

Output columns:
  year, month, category, subcategory, line_item, detail, value

Run:
  python3 parse_cor_xlsx.py COR-Summary_Monthly_1986-2024.xlsx cor_data.csv
  python3 parse_cor_xlsx.py COR-Summary_Monthly_1986-2024.xlsx cor_data.csv 2024
"""

import csv
import sys
from pathlib import Path

import openpyxl

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def find_header_row(ws):
    """Find the row containing 'Particulars' header."""
    for row_idx in range(1, 20):
        # Check columns 1-5 for 'Particulars' (layout varies by year)
        for col_idx in range(1, 6):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and "Particulars" in str(val):
                return row_idx
    return None


# Map variant month names to canonical names
MONTH_ALIASES = {
    "Jan": "Jan", "Feb": "Feb", "Mar": "Mar", "Apr": "Apr",
    "May": "May", "Jun": "Jun", "June": "Jun",
    "Jul": "Jul", "July": "Jul",
    "Aug": "Aug", "Sep": "Sep", "Sept": "Sep",
    "Oct": "Oct", "Nov": "Nov", "Dec": "Dec",
}


def get_month_columns(ws, header_row):
    """Return dict mapping canonical month name -> column index."""
    month_cols = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val:
            cleaned = str(val).strip()
            canonical = MONTH_ALIASES.get(cleaned)
            if canonical:
                month_cols[canonical] = col_idx
    return month_cols


def extract_sheet_data(ws, year):
    """Extract all data rows from a worksheet."""
    header_row = find_header_row(ws)
    if header_row is None:
        return []

    month_cols = get_month_columns(ws, header_row)
    if not month_cols:
        return []

    # Detect layout based on where months start:
    # 2004+: months at col 6+, hierarchy in cols 2-5
    # 2001-2003: months at col 4-5, hierarchy in cols 2-3  
    # pre-2001: months at col 4, hierarchy in cols 1-3
    min_month_col = min(month_cols.values())
    
    # Determine which layout
    if min_month_col >= 6:
        layout = 'new'  # 2004+
    elif min_month_col >= 4:
        layout = 'transition'  # 2001-2003
    else:
        layout = 'old'  # shouldn't happen but fallback
    
    rows = []
    # Track current hierarchy
    hierarchy = {1: "", 2: "", 3: "", 4: "", 5: ""}

    for row_idx in range(header_row + 1, ws.max_row + 1):
        # Find which column has the label
        label = None
        label_col = None
        if layout == 'new':
            max_label_col = 6  # check cols 1-5
        else:  # transition or old
            max_label_col = 4  # check cols 1-3
        
        for col in range(1, max_label_col):
            val = ws.cell(row=row_idx, column=col).value
            if val is not None and str(val).strip():
                label = str(val).strip()
                label_col = col
                break

        if label is None:
            continue

        # Stop at Notes section
        if label == "Notes:":
            break

        # Skip "of which:" helper labels
        if "of which" in label.lower():
            continue

        # Update hierarchy
        # Clear all levels >= current level
        for lvl in range(label_col, 6):
            hierarchy[lvl] = ""
        hierarchy[label_col] = label

        # Check if this row has numeric data
        has_data = False
        for m in MONTHS:
            col_idx = month_cols.get(m)
            if col_idx:
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None and isinstance(val, (int, float)):
                    has_data = True
                    break

        if not has_data:
            continue

        # Build hierarchy fields based on layout
        if layout == 'new':
            # 2004+: Col 1-2=category, Col 3=subcategory, Col 4=line_item, Col 5=detail
            category = hierarchy[1] or hierarchy[2]
            subcategory = hierarchy[3]
            line_item_base = hierarchy[4]
            detail = hierarchy[5]

            if label_col <= 2:
                category = label
                subcategory = ""
                line_item_base = ""
                detail = ""
            elif label_col == 3:
                subcategory = label
                line_item_base = ""
                detail = ""
            elif label_col == 4:
                line_item_base = label
                detail = ""
            elif label_col == 5:
                detail = label
        else:
            # transition (2001-2003) and old (pre-2001): Col 1-2=category, Col 3=line_item
            # (both use same structure, just different starting columns for months)
            category = hierarchy[1] or hierarchy[2]
            subcategory = hierarchy[2] if hierarchy[1] else ""
            line_item_base = hierarchy[3]
            detail = ""
            
            if label_col == 1:
                category = label
                subcategory = ""
                line_item_base = ""
            elif label_col == 2:
                subcategory = label
                line_item_base = ""
            elif label_col == 3:
                line_item_base = label

        # Combine line_item and detail: "BIR - Documentary Stamp"
        if detail:
            line_item = f"{line_item_base} - {detail}".strip(" -")
        else:
            line_item = line_item_base

        # Extract month values
        for m in MONTHS:
            col_idx = month_cols.get(m)
            if not col_idx:
                continue
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            # Convert to number
            if isinstance(val, (int, float)):
                numeric_val = val
            else:
                try:
                    numeric_val = float(str(val).replace(",", ""))
                except ValueError:
                    continue

            rows.append({
                "year": year,
                "month": m,
                "category": category,
                "subcategory": subcategory,
                "line_item": line_item,
                "value": numeric_val,
            })

    return rows


def parse_workbook(xlsx_path, only_year=None):
    """Parse all sheets (or just one year) from the workbook."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    all_rows = []

    for sheet_name in wb.sheetnames:
        year = sheet_name.strip()
        if only_year and year != only_year:
            continue

        ws = wb[sheet_name]
        rows = extract_sheet_data(ws, year)
        all_rows.extend(rows)
        print(f"Extracted {len(rows)} rows from sheet '{sheet_name}'", file=sys.stderr)

    return all_rows


def main():
    if len(sys.argv) < 3:
        print("Usage: python3 parse_cor_xlsx.py INPUT.xlsx OUTPUT.csv [YEAR]", file=sys.stderr)
        sys.exit(1)

    xlsx_path = Path(sys.argv[1])
    out_csv = Path(sys.argv[2])
    only_year = sys.argv[3] if len(sys.argv) > 3 else None

    rows = parse_workbook(xlsx_path, only_year)

    if not rows:
        print("No data extracted.", file=sys.stderr)
        sys.exit(1)

    fieldnames = ["year", "month", "category", "subcategory", "line_item", "value"]

    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    print(f"Wrote {len(rows)} rows to {out_csv}")


if __name__ == "__main__":
    main()
